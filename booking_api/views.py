import json
import csv
from datetime import date
from io import TextIOWrapper
from pathlib import Path

from django.conf import settings
from openpyxl import load_workbook

from django.http import JsonResponse, FileResponse
from django.shortcuts import render
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.http import require_http_methods

from Services.Bookinglogic import (
	BookingConflictError,
	cancel_booking,
	cancel_booking_from_settings,
	clear_bookings,
	create_booking,
	list_bookings,
)
from Services.count import get_system_counts
from Services.form import BookingValidationError
from Services.reccomendation import recommend_supervisor_slots
from booking_api.bootstrap import create_schedule_day_config
from booking_api.models import Booking, Panel, ScheduleDay, Slot, SupervisorStudentLink, Supervisor


@ensure_csrf_cookie
def index(request):
	return render(request, "booking_api/index.html")


@ensure_csrf_cookie
def load_data_page(request):
	return render(request, "booking_api/load_data.html")


@require_http_methods(["GET"])
def schedule_config(request):
	payload = []
	days = ScheduleDay.objects.prefetch_related("panels", "slots").all()
	for day in days:
		day_panels = [panel.name for panel in day.panels.all()]
		student_slots = [
			slot.label
			for slot in day.slots.all()
			if slot.role == Slot.ROLE_STUDENT
		]
		supervisor_slots = [
			slot.label
			for slot in day.slots.all()
			if slot.role == Slot.ROLE_SUPERVISOR
		]

		payload.append(
			{
				"date": day.date.isoformat(),
				"displayDate": day.date.strftime("%A %d %b"),
				"panels": day_panels,
				"studentSlots": student_slots,
				"supervisorSlots": supervisor_slots,
			}
		)

	return JsonResponse(payload, safe=False)


def _parse_json_request(request):
	try:
		return json.loads(request.body.decode("utf-8") or "{}")
	except json.JSONDecodeError as error:
		raise BookingValidationError("Invalid JSON payload.") from error


def _require_settings_password(payload: dict) -> None:
	password = (payload.get("password") or "").strip()
	if password != settings.SCHEDULE_SETTINGS_PASSWORD:
		raise BookingValidationError("Incorrect settings password.")


def _normalize_schedule_items(raw_value, field_name: str) -> list[str]:
	if isinstance(raw_value, list):
		items = raw_value
	elif isinstance(raw_value, str):
		items = raw_value.splitlines()
	else:
		items = []

	normalized = []
	seen = set()
	for item in items:
		value = str(item or "").strip()
		key = value.lower()
		if not value or key in seen:
			continue
		seen.add(key)
		normalized.append(value)

	if not normalized:
		raise BookingValidationError(f"Add at least one {field_name}.")

	return normalized


@require_http_methods(["POST"])
def unlock_settings(request):
	try:
		payload = _parse_json_request(request)
		_require_settings_password(payload)
	except BookingValidationError as error:
		return JsonResponse({"message": str(error)}, status=400)

	return JsonResponse({"ok": True})


@require_http_methods(["POST", "DELETE"])
def schedule_days(request):
	try:
		payload = _parse_json_request(request)
		_require_settings_password(payload)
	except BookingValidationError as error:
		return JsonResponse({"message": str(error)}, status=400)

	date_value = (payload.get("date") or "").strip()
	if not date_value:
		return JsonResponse({"message": "Choose a date."}, status=400)

	try:
		selected_date = date.fromisoformat(date_value)
	except ValueError:
		return JsonResponse({"message": "Choose a valid date."}, status=400)

	if request.method == "POST":
		try:
			panels = _normalize_schedule_items(payload.get("panels"), "panel")
			student_slots = _normalize_schedule_items(payload.get("studentSlots"), "time slot")
		except BookingValidationError as error:
			return JsonResponse({"message": str(error)}, status=400)

		day = ScheduleDay.objects.prefetch_related("panels", "slots").filter(date=selected_date).first()
		if day is not None:
			removed_panels = set(day.panels.values_list("name", flat=True)) - set(panels)
			removed_slots = set(
				day.slots.filter(role=Slot.ROLE_STUDENT).values_list("label", flat=True)
			) - set(student_slots)

			if removed_panels and Booking.objects.filter(
				day=day,
				status=Booking.STATUS_ACTIVE,
				panel__name__in=removed_panels,
			).exists():
				return JsonResponse({"message": "Cannot remove a panel that already has bookings."}, status=400)

			if removed_slots and Booking.objects.filter(
				day=day,
				status=Booking.STATUS_ACTIVE,
				role=Slot.ROLE_STUDENT,
				slot__label__in=removed_slots,
			).exists():
				return JsonResponse({"message": "Cannot remove a time slot that already has bookings."}, status=400)

		day = create_schedule_day_config(selected_date, panels=panels, student_slots=student_slots)
		return JsonResponse(
			{
				"date": day.date.isoformat(),
				"displayDate": day.date.strftime("%A %d %b"),
				"panels": panels,
				"studentSlots": student_slots,
			},
			status=201,
		)

	day = ScheduleDay.objects.filter(date=selected_date).first()
	if day is None:
		return JsonResponse({"message": "That date is not configured."}, status=404)
	if Booking.objects.filter(day=day, status=Booking.STATUS_ACTIVE).exists():
		return JsonResponse({"message": "Remove bookings for this date before deleting it."}, status=400)

	day.delete()
	return JsonResponse({"deleted": date_value})


@require_http_methods(["GET", "POST", "DELETE"])
def bookings(request):
	if request.method == "GET":
		return JsonResponse(list_bookings(), safe=False)

	if request.method == "DELETE":
		deleted_count = clear_bookings()
		return JsonResponse({"deleted": deleted_count})

	try:
		payload = json.loads(request.body.decode("utf-8") or "{}")
	except json.JSONDecodeError:
		return JsonResponse({"message": "Invalid JSON payload."}, status=400)

	try:
		booking = create_booking(payload)
	except (BookingValidationError, BookingConflictError) as error:
		return JsonResponse({"message": str(error)}, status=400)

	return JsonResponse(booking, status=201)


@require_http_methods(["GET"])
def system_counts(request):
	return JsonResponse(get_system_counts())


def _normalize_header(value: str) -> str:
	return (value or "").strip().lower().replace(" ", "_")


def _column_value(row: dict, candidates: list[str]) -> str:
	for candidate in candidates:
		if candidate in row and row[candidate] is not None:
			return str(row[candidate]).strip()
	return ""


def _parse_csv(file_obj):
	wrapper = TextIOWrapper(file_obj, encoding="utf-8-sig")
	reader = csv.DictReader(wrapper)
	for raw_row in reader:
		yield {_normalize_header(key): value for key, value in raw_row.items()}


def _parse_xlsx(file_obj):
	workbook = load_workbook(file_obj, read_only=True, data_only=True)
	sheet = workbook.active
	headers = []
	for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
		if row_index == 0:
			headers = [_normalize_header(str(value or "")) for value in row]
			continue

		yield {
			headers[index]: ("" if value is None else str(value).strip())
			for index, value in enumerate(row)
			if index < len(headers) and headers[index]
		}


@require_http_methods(["POST"])
def upload_supervisor_links(request):
	upload = request.FILES.get("file")
	if upload is None:
		return JsonResponse({"message": "Select an Excel or CSV file to upload."}, status=400)

	file_name = upload.name.lower()
	if not (file_name.endswith(".xlsx") or file_name.endswith(".csv")):
		return JsonResponse({"message": "Unsupported file type. Use .xlsx or .csv."}, status=400)

	if file_name.endswith(".csv"):
		rows = _parse_csv(upload.file)
	else:
		rows = _parse_xlsx(upload.file)

	supervisor_columns = ["supervisor", "supervisor_name", "supervisorname"]
	supervisor_email_columns = ["supervisor_email", "supervisoremail", "supervisor_mail"]
	student_columns = ["student", "student_name", "studentname"]
	student_email_columns = ["student_email", "studentemail", "student_mail"]

	pairs = []
	skipped = 0
	for row in rows:
		supervisor_name = _column_value(row, supervisor_columns)
		student_name = _column_value(row, student_columns)

		if not supervisor_name and not student_name:
			continue

		if not supervisor_name or not student_name:
			skipped += 1
			continue

		supervisor_email = _column_value(row, supervisor_email_columns)
		student_email = _column_value(row, student_email_columns)

		pairs.append((supervisor_name, supervisor_email, student_name, student_email))

	if not pairs:
		return JsonResponse(
			{"message": "No valid rows found. Required columns: supervisor and student."},
			status=400,
		)

	SupervisorStudentLink.objects.all().delete()

	seen = set()
	to_create = []
	for supervisor_name, supervisor_email, student_name, student_email in pairs:
		key = (supervisor_name.strip().lower(), student_name.strip().lower())
		if key in seen:
			continue
		seen.add(key)
		to_create.append(
			SupervisorStudentLink(
				supervisor_name=supervisor_name.strip(),
				supervisor_email=supervisor_email.strip().lower(),
				student_name=student_name.strip(),
				student_email=student_email.strip().lower(),
			)
		)

	SupervisorStudentLink.objects.bulk_create(to_create)

	return JsonResponse(
		{
			"inserted": len(to_create),
			"skipped": skipped,
			"counts": get_system_counts(),
		}
	)


@require_http_methods(["GET"])
def download_template(request):
	"""Serve the supervisor-student mapping template CSV."""
	template_path = Path(__file__).resolve().parent.parent / "supervisor_student_template.csv"
	if not template_path.exists():
		return JsonResponse({"message": "Template not found."}, status=404)
	return FileResponse(
		open(template_path, "rb"),
		as_attachment=True,
		filename="supervisor_student_template.csv",
		content_type="text/csv",
	)


@require_http_methods(["POST"])
def cancel_booking_view(request, booking_id):
	try:
		body = json.loads(request.body.decode("utf-8") or "{}")
	except json.JSONDecodeError:
		return JsonResponse({"message": "Invalid JSON."}, status=400)

	email = (body.get("email") or "").strip()
	reason = (body.get("reason") or "").strip()

	if not email:
		return JsonResponse({"message": "Email is required to cancel."}, status=400)

	try:
		updated = cancel_booking(booking_id, email, reason)
	except BookingValidationError as error:
		return JsonResponse({"message": str(error)}, status=400)
	except BookingConflictError as error:
		return JsonResponse({"message": str(error)}, status=403)

	return JsonResponse(updated)


@require_http_methods(["POST"])
def admin_cancel_booking_view(request, booking_id):
	try:
		payload = _parse_json_request(request)
		_require_settings_password(payload)
	except BookingValidationError as error:
		return JsonResponse({"message": str(error)}, status=400)

	try:
		updated = cancel_booking_from_settings(booking_id, payload.get("reason") or "")
	except BookingValidationError as error:
		return JsonResponse({"message": str(error)}, status=400)

	return JsonResponse(updated)


@require_http_methods(["GET"])
def recommendations(request):
	supervisor_name = (request.GET.get("supervisor") or "").strip()
	date_value = (request.GET.get("date") or "").strip()
	panel_name = (request.GET.get("panel") or "").strip()
	if not supervisor_name or not date_value or not panel_name:
		return JsonResponse({"recommendations": []})

	day = ScheduleDay.objects.filter(date=date_value).first()
	panel = Panel.objects.filter(day=day, name=panel_name).first() if day else None
	if not day or not panel:
		return JsonResponse({"recommendations": []})

	slots = Slot.objects.filter(day=day, role=Slot.ROLE_SUPERVISOR).order_by("sort_order", "label")
	labels = recommend_supervisor_slots(supervisor_name, slots, day, panel)
	return JsonResponse({"recommendations": labels})


@require_http_methods(["GET"])
def search_supervisors(request):
	"""Search supervisors by name prefix. Returns list of {name, email}."""
	query = (request.GET.get("q") or "").strip().lower()
	if not query or len(query) < 2:
		return JsonResponse({"results": []})

	supervisors = Supervisor.objects.filter(name__icontains=query).order_by("name")[:20]
	results = [{"name": s.name, "email": s.email} for s in supervisors]
	return JsonResponse({"results": results})


@require_http_methods(["GET"])
def export_bookings(request):
	"""Export all active bookings as CSV."""
	from django.http import HttpResponse
	from Services.count import serialize_booking
	from booking_api.models import Booking

	bookings = Booking.objects.filter(status=Booking.STATUS_ACTIVE).order_by("-booked_at")
	
	response = HttpResponse(content_type="text/csv")
	response["Content-Disposition"] = 'attachment; filename="bookings_export.csv"'
	
	writer = csv.writer(response)
	writer.writerow(["First Name", "Surname", "Email", "Role", "Supervisor", "Co-Supervisor", "Date", "Panel", "Slot", "Booked At", "Status"])
	
	for booking in bookings:
		writer.writerow([
			booking.first_name,
			booking.surname,
			booking.email,
			booking.role,
			booking.supervisor,
			booking.co_supervisor,
			booking.day.date.isoformat(),
			booking.panel.name,
			booking.slot.label,
			booking.booked_at.isoformat(),
			booking.status,
		])
	
	return response
