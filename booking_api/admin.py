import csv
from datetime import date
from io import TextIOWrapper

from django.contrib import admin, messages
from django import forms
from django.shortcuts import redirect
from django.template.response import TemplateResponse
from django.urls import path
from openpyxl import load_workbook

from booking_api.bootstrap import DEFAULT_BULK_STUDENT_SLOTS, create_schedule_day_config
from booking_api.models import Booking, Panel, ScheduleDay, Slot, SupervisorStudentLink, Supervisor


def _normalize_header(value):
    return (value or "").strip().lower().replace(" ", "_")


def _column_value(row, candidates):
    for key in candidates:
        if key in row and row[key] is not None and str(row[key]).strip():
            return str(row[key]).strip()
    return ""


def _rows_from_xlsx(file_obj):
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    sheet = workbook.active
    headers = []
    for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
        if row_index == 0:
            headers = [_normalize_header(str(v or "")) for v in row]
            continue
        yield {
            headers[i]: ("" if v is None else str(v).strip())
            for i, v in enumerate(row)
            if i < len(headers) and headers[i]
        }


def _rows_from_csv(file_obj):
    wrapper = TextIOWrapper(file_obj, encoding="utf-8-sig")
    reader = csv.DictReader(wrapper)
    for raw in reader:
        yield {_normalize_header(k): v for k, v in raw.items()}


class BulkScheduleSeedForm(forms.Form):
    booking_type = forms.ChoiceField(
        label="Booking type",
        choices=ScheduleDay.BOOKING_TYPE_CHOICES,
        initial=ScheduleDay.BOOKING_TYPE_SYNDICATE,
        help_text="Dates will be created only for this page type.",
    )
    panel_name = forms.CharField(
        label="Panel name",
        max_length=100,
        initial="Panel 1",
        help_text="This panel name will be created for every selected date.",
    )
    dates = forms.CharField(
        label="Dates",
        widget=forms.Textarea(attrs={"rows": 8, "placeholder": "2026-08-03\n2026-08-04\n2026-08-05"}),
        help_text="Enter one ISO date per line, or separate multiple dates with commas.",
    )

    def clean_dates(self):
        raw_value = self.cleaned_data["dates"]
        tokens = raw_value.replace(",", "\n").splitlines()
        parsed_dates = []
        seen = set()

        for token in tokens:
            value = token.strip()
            if not value:
                continue
            try:
                parsed_date = date.fromisoformat(value)
            except ValueError as error:
                raise forms.ValidationError(f"Use YYYY-MM-DD format. Invalid value: {value}") from error

            if parsed_date < date.today():
                raise forms.ValidationError(f"Past dates are not allowed here: {value}")

            if parsed_date in seen:
                continue

            seen.add(parsed_date)
            parsed_dates.append(parsed_date)

        if not parsed_dates:
            raise forms.ValidationError("Add at least one valid date.")

        return parsed_dates


@admin.register(ScheduleDay)
class ScheduleDayAdmin(admin.ModelAdmin):
    list_display = ("id", "date", "booking_type")
    list_filter = ("booking_type",)
    changelist_template = "admin/booking_api/scheduleday/change_list.html"

    def get_urls(self):
        urls = super().get_urls()
        extra = [
            path(
                "bulk-seed/",
                self.admin_site.admin_view(self.bulk_seed_view),
                name="booking_api_scheduleday_bulk_seed",
            ),
        ]
        return extra + urls

    def bulk_seed_view(self, request):
        form = BulkScheduleSeedForm(request.POST or None)

        if request.method == "POST" and form.is_valid():
            booking_type = form.cleaned_data["booking_type"]
            panel_name = form.cleaned_data["panel_name"].strip()
            selected_dates = form.cleaned_data["dates"]
            created_count = 0
            updated_count = 0
            skipped_dates = []

            for selected_date in selected_dates:
                existing_day = ScheduleDay.objects.filter(date=selected_date, booking_type=booking_type).first()
                if existing_day and Booking.objects.filter(
                    day=existing_day,
                    status=Booking.STATUS_ACTIVE,
                ).exists():
                    skipped_dates.append(selected_date.isoformat())
                    continue

                create_schedule_day_config(
                    selected_date,
                    booking_type=booking_type,
                    panels=[panel_name],
                    student_slots=DEFAULT_BULK_STUDENT_SLOTS,
                )

                if existing_day is None:
                    created_count += 1
                else:
                    updated_count += 1

            if created_count or updated_count:
                summary = (
                    f"Bulk seed complete: {created_count} created, {updated_count} updated. "
                    f"Each {booking_type} date now has panel '{panel_name}' and half-hour slots from 09:00 to 18:00."
                )
                self.message_user(request, summary, level=messages.SUCCESS)

            if skipped_dates:
                self.message_user(
                    request,
                    "Skipped dates with active bookings: " + ", ".join(skipped_dates),
                    level=messages.WARNING,
                )

            if not created_count and not updated_count and skipped_dates:
                return redirect("..")

            return redirect("..")

        context = {
            **self.admin_site.each_context(request),
            "title": "Bulk Seed Schedule Dates",
            "opts": self.model._meta,
            "form": form,
            "slot_range": f"{DEFAULT_BULK_STUDENT_SLOTS[0]} to {DEFAULT_BULK_STUDENT_SLOTS[-1]}",
        }
        return TemplateResponse(
            request,
            "admin/booking_api/scheduleday/bulk_seed.html",
            context,
        )


@admin.register(Panel)
class PanelAdmin(admin.ModelAdmin):
    list_display = ("id", "day", "name", "sort_order")
    list_filter = ("day__booking_type", "day",)


@admin.register(Slot)
class SlotAdmin(admin.ModelAdmin):
    list_display = ("id", "day", "role", "label", "sort_order")
    list_filter = ("day__booking_type", "day", "role")


@admin.register(SupervisorStudentLink)
class SupervisorStudentLinkAdmin(admin.ModelAdmin):
    list_display = ("id", "supervisor_name", "supervisor_email", "student_name", "student_email")
    search_fields = ("supervisor_name", "supervisor_email", "student_name", "student_email")
    changelist_template = "admin/booking_api/supervisorstudentlink/change_list.html"

    def get_urls(self):
        urls = super().get_urls()
        extra = [
            path(
                "upload-excel/",
                self.admin_site.admin_view(self.upload_excel_view),
                name="booking_api_supervisorstudentlink_upload_excel",
            ),
        ]
        return extra + urls

    def upload_excel_view(self, request):
        if request.method == "POST":
            upload = request.FILES.get("file")
            if not upload:
                self.message_user(request, "No file selected.", level=messages.ERROR)
                return redirect("..")

            name = upload.name.lower()
            if not (name.endswith(".xlsx") or name.endswith(".csv")):
                self.message_user(request, "Use .xlsx or .csv only.", level=messages.ERROR)
                return redirect("..")

            rows = list(_rows_from_xlsx(upload.file) if name.endswith(".xlsx") else _rows_from_csv(upload.file))

            sup_cols = ["supervisor", "supervisor_name", "supervisorname"]
            stu_cols = ["student", "student_name", "studentname"]

            seen, to_create, skipped = set(), [], 0
            for row in rows:
                sup = _column_value(row, sup_cols)
                stu = _column_value(row, stu_cols)
                if not sup and not stu:
                    continue
                if not sup or not stu:
                    skipped += 1
                    continue
                key = (sup.lower(), stu.lower())
                if key in seen:
                    continue
                seen.add(key)
                to_create.append(SupervisorStudentLink(supervisor_name=sup, student_name=stu))

            if not to_create:
                self.message_user(
                    request,
                    "No valid rows found. Required columns: supervisor and student.",
                    level=messages.ERROR,
                )
                return redirect("..")

            SupervisorStudentLink.objects.all().delete()
            SupervisorStudentLink.objects.bulk_create(to_create)

            msg = f"Import complete: {len(to_create)} mappings loaded."
            if skipped:
                msg += f" {skipped} incomplete rows skipped."
            self.message_user(request, msg, level=messages.SUCCESS)
            return redirect("..")

        context = {
            **self.admin_site.each_context(request),
            "title": "Upload Supervisor-Student Mapping",
            "opts": self.model._meta,
        }
        return TemplateResponse(
            request,
            "admin/booking_api/supervisorstudentlink/upload_excel.html",
            context,
        )


@admin.register(Supervisor)
class SupervisorAdmin(admin.ModelAdmin):
	list_display = ("id", "name", "email")
	search_fields = ("name", "email")
	ordering = ("name",)


@admin.register(Booking)
class BookingAdmin(admin.ModelAdmin):
	list_display = ("id", "first_name", "surname", "email", "booking_type", "role", "supervisor", "get_co_supervisor", "day", "panel", "slot", "status", "booked_at")
	list_filter = ("booking_type", "day", "role", "panel", "status")
	search_fields = ("first_name", "surname", "email", "supervisor", "co_supervisor")
	readonly_fields = ("booked_at", "cancelled_at")

	def get_co_supervisor(self, obj):
		return obj.co_supervisor if hasattr(obj, 'co_supervisor') else ""
	get_co_supervisor.short_description = "Co-Supervisor"
